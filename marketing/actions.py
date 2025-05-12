import csv
import json
from datetime import datetime
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _

def export_as_csv(modeladmin, request, queryset):
    """
    Export selected items as CSV
    """
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ('password',)]
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.lower()}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    writer = csv.writer(response)
    
    writer.writerow(field_names)
    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if callable(value):
                value = value()
            row.append(str(value))
        writer.writerow(row)
    
    return response

export_as_csv.short_description = _("Export selected items as CSV")

def export_as_json(modeladmin, request, queryset):
    """
    Export selected items as JSON
    """
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ('password',)]
    
    response = HttpResponse(content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.lower()}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    
    data = []
    for obj in queryset:
        item = {}
        for field in field_names:
            value = getattr(obj, field)
            if callable(value):
                value = value()
            
            # Handle datetime objects for JSON serialization
            if isinstance(value, datetime):
                value = value.isoformat()
            
            item[field] = value
        data.append(item)
    
    response.write(json.dumps(data, default=str, indent=4))
    return response

export_as_json.short_description = _("Export selected items as JSON")

def export_as_excel(modeladmin, request, queryset):
    """
    Export selected items as Excel (requires openpyxl)
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        modeladmin.message_user(request, "The openpyxl package is required for Excel export", level='error')
        return
    
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ('password',)]
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.lower()}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = meta.verbose_name_plural[:31]  # Max worksheet title length in Excel
    
    # Write headers
    for col_num, column_title in enumerate(field_names, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = openpyxl.styles.Font(bold=True)
        
        # Set column width based on content
        worksheet.column_dimensions[get_column_letter(col_num)].width = max(15, len(column_title) + 2)
    
    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(field_names, 1):
            value = getattr(obj, field)
            if callable(value):
                value = value()
            worksheet.cell(row=row_num, column=col_num).value = value
    
    workbook.save(response)
    return response

export_as_excel.short_description = _("Export selected items as Excel") 